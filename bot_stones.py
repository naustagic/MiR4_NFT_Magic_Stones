import discord
import requests
import json
import os
from dotenv import load_dotenv
from collections import deque
from typing import Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

TOKEN = os.getenv("TOKEN")
CHANNEL_ID = int(os.getenv("CHANNEL_ID"))

intents = discord.Intents.default()
intents.message_content = True  # Permite ao bot ler o conteúdo das mensagens

client = discord.Client(intents=intents)

# Fila de solicitações
request_queue = deque()
processing_request: Optional[discord.Message] = None

async def update_status(message: discord.Message, status: str):
    # Atualiza a mensagem de status com o progresso
    await message.edit(content=status)

async def process_request(message: discord.Message):
    global processing_request

    if processing_request:
        await message.reply("Por favor, aguarde a conclusão da solicitação anterior antes de fazer uma nova.")
        return

    processing_request = message

    if len(request_queue) >= 3:
        position = len(request_queue) - 2
        await message.reply(f"A fila está cheia. Você está na posição {position + 1} da fila de solicitações.")
        request_queue.append(message)
        return

    status_message = await message.reply("Por favor, aguarde enquanto o XLSX está sendo gerado...")

    try:
        nft_id = message.content.split('/')[-1]
        
        # URL da API para obter o transportID
        summary_url = f"https://webapi.mir4global.com/nft/character/summary?seq={nft_id}&languageCode=en"
        summary_response = requests.get(summary_url)
        await update_status(status_message, "Obtendo dados do resumo...")

        if summary_response.status_code == 200:
            summary_data = summary_response.json()
            
            if 'data' in summary_data and 'character' in summary_data['data']:
                transport_id = summary_data['data']['character']['transportID']
                
                inven_url = f"https://webapi.mir4global.com/nft/character/inven?transportID={transport_id}&languageCode=en"
                inven_response = requests.get(inven_url)
                await update_status(status_message, "Obtendo dados do inventário...")

                if inven_response.status_code == 200:
                    data = inven_response.json()
                    
                    if 'data' in data:
                        items = data['data']
                        
                        filtered_items = [
                            {"itemName": item.get("itemName", "N/A"), "itemUID": item.get("itemUID", "N/A")}
                            for item in items
                            if item.get("itemName", "").startswith("[L] Magic Stone") or item.get("itemName", "").startswith("[E] Magic Stone")
                        ]
                        
                        item_details = []
                        base_url = "https://webapi.mir4global.com/nft/character/itemdetail"

                        total_items = len(filtered_items)
                        for index, item in enumerate(filtered_items):
                            itemUID = item["itemUID"]
                            detail_url = f"{base_url}?transportID={transport_id}&class=1&itemUID={itemUID}&languageCode=en"
                            response = requests.get(detail_url)
                            
                            await update_status(status_message, f"Processando Magic Stones {index + 1} de {total_items}...")

                            if response.status_code == 200:
                                item_detail = response.json()
                                item_info = {
                                    "itemName": item_detail["data"]["itemName"],
                                    "options": []
                                }
                                
                                def calculate_final_value(option_value, trance_step, trance_value, format):
                                    final_value = option_value + (trance_step or 0) + trance_value
                                    return round(final_value, 2)

                                option_sums = {}
                                
                                for option in item_detail["data"]["options"]:
                                    option_name = option["optionName"]
                                    option_value = calculate_final_value(option["optionValue"], option.get("tranceValue"), 0, option["optionFormat"])
                                    if option_name in option_sums:
                                        option_sums[option_name] += option_value
                                    else:
                                        option_sums[option_name] = option_value
                                
                                for add_option in item_detail["data"]["addOptions"]:
                                    option_name = add_option["optionName"]
                                    option_value = calculate_final_value(add_option["optionValue"], add_option.get("optionTranceStep"), add_option.get("tranceValue", 0), add_option["optionAddFormat"])
                                    if option_name in option_sums:
                                        option_sums[option_name] += option_value
                                    else:
                                        option_sums[option_name] = option_value
                                
                                for option_name, option_value in option_sums.items():
                                    option_format = next((opt["optionAddFormat"] for opt in item_detail["data"]["addOptions"] if opt["optionName"] == option_name), "")
                                    item_info["options"].append({
                                        "optionName": option_name,
                                        "optionValue": f'{option_value}%' if option_format == "%" else option_value
                                    })
                                
                                item_details.append(item_info)
                            else:
                                print(f"Falha ao acessar a API para itemUID {itemUID}. Status code: {response.status_code}")

                        # Criar um arquivo XLSX
                        wb = Workbook()
                        ws = wb.active
                        headers = ["Stone Name"] + list({option["optionName"] for item in item_details for option in item["options"]})
                        ws.append(headers)

                        for item in item_details:
                            row = [item["itemName"]]
                            option_dict = {option["optionName"]: option["optionValue"] for option in item["options"]}
                            row.extend(option_dict.get(option_name, "") for option_name in headers[1:])  # Ignorando o primeiro item (Stone Name)
                            ws.append(row)

                        # Formatar células
                        for row in range(2, len(item_details) + 2):  # Começa da linha 2
                            for col in range(2, len(headers) + 1):  # Começa da coluna 2
                                cell_value = ws.cell(row=row, column=col).value
                                if isinstance(cell_value, str) and cell_value.endswith('%'):
                                    # Converte a string de porcentagem para número decimal
                                    percentage_value = float(cell_value[:-1].replace(',', '.')) / 100
                                    ws.cell(row=row, column=col).value = percentage_value
                                    ws.cell(row=row, column=col).number_format = '0.0%'  # Formato de porcentagem com uma casa decimal
                                elif isinstance(cell_value, (int, float)):
                                    ws.cell(row=row, column=col).number_format = '0'  # Formato numérico sem casas decimais

                        xlsx_filename = f"legendary_stones_{nft_id}.xlsx"
                        wb.save(xlsx_filename)

                        await update_status(status_message, f"Seu arquivo XLSX está pronto para download: {xlsx_filename}")
                        await message.reply(file=discord.File(xlsx_filename))

                        # Limpeza de arquivos temporários
                        os.remove(xlsx_filename)
                    
                    else:
                        await update_status(status_message, "A chave 'data' não foi encontrada na resposta da API de inventário.")
                else:
                    await update_status(status_message, f"Falha ao acessar a API de inventário. Status code: {inven_response.status_code}")
            else:
                await update_status(status_message, "A chave 'data' ou 'character' não foi encontrada na resposta da API de resumo.")
        else:
            await update_status(status_message, f"Falha ao acessar a API de resumo. Status code: {summary_response.status_code}")
    except Exception as e:
        await update_status(status_message, f"Ocorreu um erro: {str(e)}")
    finally:
        processing_request = None
        if request_queue:
            next_request = request_queue.popleft()
            await process_request(next_request)

@client.event
async def on_message(message):
    if message.channel.id != CHANNEL_ID or message.author == client.user:
        return

    # Verifica se a URL começa com "https://" e contém "xdraco.com/nft/trade/"
    if message.content.startswith('https://') and 'xdraco.com/nft/trade/' in message.content:
        if len(message.content.strip()) <= len('https://xdraco.com/nft/trade/'):
            await message.reply("Link inválido. Por favor, envie um link completo no formato Ex: 'https://xdraco.com/nft/trade/11111'.")
        else:
            await process_request(message)

client.run(TOKEN)
